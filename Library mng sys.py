import pandas as pd
from openpyxl import Workbook, load_workbook

excel_file = "Library_data.xlsx"
workbook = load_workbook(excel_file)
sheet_1 = workbook["Books_data"]
sheet_2 = workbook["Issued_data"]
sheet_3 = workbook["Returned_data"]
sheet_4 = workbook["Admin"]
class Admin:
    def __init__(self):
        self.current_admin_row = None

    def validate_admin(self) -> bool:
        Admin_ID = input("ENTER ADMIN ID:       ")
        Password = input("ENTER THE PASSWORD:   ")
        for row in range(2, sheet_4.max_row + 1):
            admin_name = sheet_4.cell(row=row, column=1).value
            password = sheet_4.cell(row=row, column=2).value
            if Admin_ID == admin_name and Password == password:
                print("ACCESS GRANTED!")
                self.current_admin_row = row
                return True
        print("WRONG ADMIN ID OR PASSWORD")
        return False
    
    # FUNCTION TO CHANGE PASSWORD FOR ADMIN    
    def change_password(self) -> bool:
        if self.validate_admin():
            new_pass = input("ENTER NEW PASSWORD:   ")
            sheet_4.cell(row=self.current_admin_row, column=2, value=new_pass)
            workbook.save(excel_file)
            print("PASSWORD CHANGED SUCCESSFULLY!")
            return True
        return False

class DashBoard:
    book_ID = None
    book_name = None
    book_type = None
    book_status = None
    student_cnic = None
    issue_date = None
    available = None

    # FUNCTION TO SEARCH BOOK FROM THE MAIN DATA OF LIBRARY
    def search_book(self, ID) -> bool:
        for row in range(2, sheet_1.max_row + 1):
            current_ID = sheet_1.cell(row=row, column=1).value
            if current_ID == ID:
                return True
        return False

    # FUNCTION TO GET NAME OF BOOK FROM BOOK'S MAIN DATA
    def get_book_name(self,ID)->str:
        for row in range(2,sheet_1.max_row+1):
            book = sheet_1.cell(row=row, column=1).value
            if(ID == book):
                name = sheet_1.cell(row=row,column=2).value 
                return name
        return ""
    
    # FUNCTION TO ADD BOOK TO THE DATA
    def add_book(self):
        n = int(input("ENTER THE NUMBER OF BOOKS ADD:    "))
        next_row = sheet_1.max_row + 1
        isSame_ID = False
        for i in range(n):
            book_ID = input(f"ENTER ID OF BOOK {i + 1}: ")
            if self.search_book(book_ID):
                print(f"THE BOOK WITH ID {book_ID} IS ALREADY IN DATA")
                isSame_ID = True
            else:
                isSame_ID = False
            while isSame_ID:
                book_ID = input(f"RE-ENTER ID OF BOOK {i+1}:    ")
                if not self.search_book(book_ID):
                    isSame_ID = False
            book_name = input(f"ENTER NAME OF BOOK {i + 1}: ")
            book_type = input(f"ENTER TYPE OF BOOK {i + 1}: ")
            book_status = "Available"
            status = int(input("PRESS\n1:   BOOK IS AVAILABLE\n2:   BOOK IS NOT AVAILABLE:  "))
            while status != 1 and status != 2:
                status = int(input("ENTER 1 or 2:   "))
            book_status = "Not Available" if status == 2 else "Available"
            sheet_1.cell(row=next_row, column=1, value=book_ID)
            sheet_1.cell(row=next_row, column=2, value=book_name)
            sheet_1.cell(row=next_row, column=3, value=book_type)
            sheet_1.cell(row=next_row, column=4, value=book_status)
            next_row += 1
        workbook.save(excel_file)

    # FUNCTION TO CHECK THE LIMIT OF BOOK TAKER TO PREVENT EXCEDDING LIMIT
    def limit_in_data(self,ID):
        cnt_issue = 0
        cnt_rtrn = 0
        for row in range(2,sheet_2.max_row+1):
            ID_in_issue = sheet_2.cell(row=row,column=1).value
            if(ID_in_issue == ID):
                cnt_issue+=1  
        for row in range(2,sheet_3.max_row+1):
            ID_in_rtrn= sheet_3.cell(row=row,column=2).value
            if(ID_in_rtrn == ID):
                cnt_rtrn+=1
        
        return max(cnt_issue-cnt_rtrn,0)
          
    # FUNCTION TO CHECK THE LIMIT OF BOOK TAKER TO PREVENT NOT CROSSING THE LIMIT
    def issue_limit(self, ID) -> int:
        cnt_issue = 0
        cnt_rtrn = 0
        for row in range(2,sheet_2.max_row+1):
            ID_in_issue = sheet_2.cell(row=row,column=1).value
            if(ID_in_issue == ID):
                cnt_issue+=1  
        for row in range(2,sheet_3.max_row+1):
            ID_in_rtrn= sheet_3.cell(row=row,column=2).value
            if(ID_in_rtrn == ID):
                cnt_rtrn+=1
        
        return max(cnt_issue-cnt_rtrn,0)

    # FUNCTION TO EDIT BOOK's INFORMATION
    def edit_book(self):
        book_ID = input("ENTER ID OF BOOK: ")
        if not self.search_book(book_ID):
            print(f"BOOK WITH ID {book_ID} NOT FOUND IN DATA")
            input("PRESS ANY KEY TO CONTINUE ")
        else:
            for i in range(2, sheet_1.max_row + 1):
                book = sheet_1.cell(row=i, column=1).value
                if book_ID == book:
                    print("")
                    print(f"BOOK ID:        {sheet_1.cell(row=i,column=1).value}")
                    print(f"BOOK NAME:      {sheet_1.cell(row=i,column=2).value}")
                    print(f"BOOK TYPE:      {sheet_1.cell(row=i,column=3).value}")
                    print(f"AVAILABILITY:   {sheet_1.cell(row=i,column=4).value.upper()}")
        new_book_name = input("ENTER NEW NAME FOR THE BOOK: ")
        new_book_type = input("ENTER NEW TYPE FOR THE BOOK: ")
        for row in range(2, sheet_1.max_row + 1):
            current_ID = sheet_1.cell(row=row, column=1).value
            if current_ID == book_ID:
                sheet_1.cell(row=row, column=2, value=new_book_name)
                sheet_1.cell(row=row, column=3, value=new_book_type)
                workbook.save(excel_file)
                print(f"BOOK WITH ID {book_ID} HAS BEEN UPDATED SUCCESSFULLY!")
                return

    # FUNCTION TO CHECK IF THE BOOK IS AVAILABLE, IF NOT THEN THIS MEANS BOOK WAS ISSUED TO SOMEONE
    def search_book_availability(self,ID)->bool:
        for row in range(2,sheet_1.max_row+1):
            book = sheet_1.cell(row=row, column=1).value
            if(ID == book):
                status = sheet_1.cell(row=row,column=4).value 
                return status == "AVAILABLE"
        return False
    
    # FUNCTION TO ISSUE BOOK
    def issue_book(self):
        book_ID = input("ENTER THE BOOK ID: ")
        if not self.search_book_availability(book_ID):
            print("BOOK IS ALREADY ISSUED")
            input("ENTER ANY KEY TO CONTINUE ")
            return
        if not self.search_book(book_ID):
            print(f"BOOK WITH ID {book_ID} NOT FOUND")
            input("ENTER ANY KEY TO CONTINUE")
            return
        name = self.get_book_name(book_ID)
        print(f"BOOK NAME:  {name}")
        student_cnic = input("ENTER THE CNIC OF STUDENT/BOOK TAKER: ")
        current_limit = 0
        current_limit = self.issue_limit(student_cnic)+1
        while (current_limit <= 3):       
            issue_date = input("ENTER ISSUE DATE:   ")
            next_row = sheet_2.max_row + 1
            sheet_2.cell(row=next_row, column=1, value=student_cnic)
            sheet_2.cell(row=next_row, column=2, value=book_ID)
            sheet_2.cell(row=next_row, column=3, value=name)
            sheet_2.cell(row=next_row, column=4, value=issue_date)
            sheet_2.cell(row=next_row, column=5, value=current_limit)
            for row in range(2, sheet_1.max_row + 1):
                current_ID = sheet_1.cell(row=row, column=1).value
                if current_ID == book_ID:
                    sheet_1.cell(row=row, column=4, value="Not Available")
                    break
            break
        workbook.save(excel_file)
        if(current_limit<=3):
            print(f"BOOK:   {name}  ID: {book_ID}   ISSUED TO {student_cnic} SUCCESSFULLY")
        else:
            print(f"LIMIT OF 3 BOOKS HAS BEEN REACHED FOR STUDENT/BOOK TAKER HAVING ID: {student_cnic}")
        input("ENTER ANY KEY TO CONTINUE ")
    
    # FUNCTION TO SEARCH/GET ISSUE DATE
    def search_issue_date(self,CNIC)->str:
        for row in range(2, sheet_2.max_row + 1):
            CNIC_in_data = sheet_2.cell(row=row, column=1).value
            if CNIC == CNIC_in_data:
                date = sheet_2.cell(row=row, column=4).value
                return date
        return ""
    
    # FUNCTION TO SEARCH IF BOOK IS AVAILABLE IN DATA MEANT IT IS NOT ISSUED OR IT HAS BEEN RETURNED
    def search_if_book_is_available(self, ID)->bool:
        if(self.search_book_availability(ID)):
            return False
        for row in range(2, sheet_1.max_row + 1):
            current_ID = sheet_1.cell(row=row, column=1).value
            if current_ID == ID:
                status = sheet_1.cell(row=row, column=4).value
                return status == "Available"
        return False
    
    # FUNCTIONS TO CHECK THE BOOK ID IN ISSUE DATA
    def search_book_in_issued_data(self,Book_ID)->bool:
        for row in range(2, sheet_2.max_row + 1):
            current_ID = sheet_2.cell(row=row, column=2).value
            if current_ID == Book_ID:
                return True
        return False
     
    # FUNCTION TO SEARCH CNIC/ID OF BOOK TAKER   
    def search_book_Taker(self,book_ID)->int:
        for row in range(2, sheet_2.max_row + 1):
            book_ID_in_data = sheet_2.cell(row=row, column=2).value
            if book_ID == book_ID_in_data:
                CNIC = sheet_2.cell(row=row, column=1).value
                return CNIC
        return 0
    
    # FUNCTIONS TO RETURN BOOK
    def return_book(self):
        next_row = sheet_3.max_row + 1
        Book_ID = input("ENTER THE BOOK ID:   ")
        if self.search_book_availability(Book_ID):
            print("AS BOOK IS AVAILABLE SO IT WAS NOT ISSUED!")
            input("PRESS ANY KEY TO CONTINUE    ")
            return
        if not self.search_book_in_issued_data(Book_ID):
            print(f"BOOK WITH ID {Book_ID} NOT FOUND IN ISSUED DATA")
            input("PRESS ANY KEY TO CONTINUE    ")
            return
        else:
            id = self.search_book_Taker(Book_ID)
            if(id != 0):                
                print(f"ID IS {id}")
                issue_date = self.search_issue_date(id)
                print(f"ISSUE DATE IS:  {issue_date}")
                Date = input("ENTER TODAY'S DATE:   ")
                charges=float(input("ENTER FINE AMOUNT (if):    "))
                sheet_3.cell(row=next_row, column=1, value=Book_ID)
                sheet_3.cell(row=next_row, column=2, value=id)
                sheet_3.cell(row=next_row, column=3, value=issue_date)
                sheet_3.cell(row=next_row, column=4, value=Date)
                sheet_3.cell(row=next_row, column=5, value=charges)
                print(f"BOOK WITH ID {Book_ID} HAS BEEN RETURNED BY ID: {id}.")
                for row in range(2, sheet_1.max_row + 1):
                    current_ID = sheet_1.cell(row=row, column=1).value
                    if current_ID == Book_ID:
                        sheet_1.cell(row=row, column=4, value="AVAILABLE")
                        break
                    # Changes limit in issued_data when book is returned
                current_limit = self.issue_limit(id)
                if current_limit > 0:
                    sheet_2.cell(row=next_row, column=5, value=current_limit)
                # else:
                #     print(f"No books currently issued for {id}")
                
                workbook.save(excel_file)
            else:
                print(f"BOOK WITH ID {Book_ID} WAS NOT ISSUED TO BOOK TAKER WITH ID {id}")
                
            workbook.save(excel_file)
            input("PRESS ANY KEY TO CONTINUE    ")    
            
    # FUNCTION TO DELETE WHOLE DATA OF PARTICULAR BOOK
    def delete_book(self):
        book_ID = input("ENTER THE ID OF THE BOOK DELETE: ")
        if not self.search_book(book_ID):
            print(f"BOOK WITH ID {book_ID} NOT FOUND IN DATA")
            input("PRESS ANY KEY TO CONTINUE ")
            return
        for row in range(2, sheet_1.max_row + 1):
            current_ID = sheet_1.cell(row=row, column=1).value
            if current_ID == book_ID:
                sheet_1.delete_rows(row, 1)
                workbook.save(excel_file)
                print(f"BOOK WITH ID {book_ID} HAS BEEN SUCCESSFULLY DELETED!")
                input("PRESS ANY KEY TO CONTINUE ")
                return
        
    # FUNCTION TO PRINT DATA OF PARTICULAR BOOK
    def show_book(self,book_ID):
        if not self.search_book(book_ID):
            print(f"BOOK WITH ID {book_ID} NOT FOUND IN DATA")
            input("PRESS ANY KEY TO CONTINUE ")
        else:
            for i in range(2, sheet_1.max_row + 1):
                book = sheet_1.cell(row=i, column=1).value
                if book_ID == book:
                    print("")
                    print(f"BOOK ID:        {sheet_1.cell(row=i,column=1).value}")
                    print(f"BOOK NAME:      {sheet_1.cell(row=i,column=2).value}")
                    print(f"BOOK TYPE:      {sheet_1.cell(row=i,column=3).value}")
                    print(f"AVAILABILITY:   {sheet_1.cell(row=i,column=4).value.upper()}")
            input("PRESS ANY KEY TO CONTINUE ")
    
    # FUNCTION TO PRINT ALL THE BOOKS IN DATA
    def show_all_books(self):
        if sheet_1.max_row < 2:
            print("No books available.")
            input("PRESS ANY KEY TO CONTINUE ")
            return
        cnt = 0
        for row in range(2, sheet_1.max_row + 1):
            book_id = sheet_1.cell(row=row, column=1).value
            book_name = sheet_1.cell(row=row, column=2).value
            book_type = sheet_1.cell(row=row, column=3).value
            book_status = sheet_1.cell(row=row,column=4).value
            if book_id is None:
                continue
            print("\n")
            print(f"BOOK ID:        {book_id}")
            print(f"BOOK NAME:      {book_name}")
            print(f"BOOK TYPE:      {book_type}")
            print(f"AVAILABILITY:   {book_status.upper()}")
            cnt+=1
        print(f"\n\nTOTAL BOOKS ARE: {cnt}")
        print("\n")
        input("PRESS ANY KEY TO CONTINUE    ")

    # FUNCTION TO LOGIN IN THE SYSTEM
    def login(self):
        print("\n\n\t\t\tWELCOME TO LIBRARY MANAGEMENT SYSTEM BY HASNAIN BASIT\n\n")
        admin_obj = Admin()
        if admin_obj.validate_admin():
            while (1):
                print("\033[H\033[J", end="") # clears screen
                choice = int(input(f"\nPRESS\n1:     CHANGE PASSWORD\n2:     ADD BOOKS\n3:     ISSUE BOOK\n4:     EDIT BOOK\n5:     RETURN BOOK\n6:     DELETE BOOK\n7:     SEARCH/SHOW BOOK\n8:     SHOW ALL BOOKS\n9:     LOGOUT\nCHOOSE:   "))
                print("\033[H\033[J", end="") # clears screen
                if(choice==1):
                    admin_obj.change_password()
                elif(choice==2):
                    self.add_book()
                    print("\033[H\033[J", end="") # clears screen
                elif(choice == 3):
                    self.issue_book()
                    print("\033[H\033[J", end="") # clears screen
                elif(choice == 4):
                    self.edit_book()
                    print("\033[H\033[J", end="") # clears screen
                elif(choice==5):
                    self.return_book()
                    print("\033[H\033[J", end="") # clears screen
                elif(choice==6):
                    self.delete_book()
                    print("\033[H\033[J", end="") # clears screen
                elif(choice ==7):
                    book_ID = input("ENTER THE ID OF BOOK:  ")
                    self.show_book(book_ID)
                    print("\033[H\033[J", end="") # clears screen
                elif(choice==8):
                    self.show_all_books()
                elif(choice==9):
                    # logout
                    print("\033[H\033[J", end="") # clears screen
                    break
                else:
                    print("ENTER VALID CHOICE BETWEEN THE GIVEN OPTIONS")

dashBoard_obj = DashBoard()
dashBoard_obj.login()